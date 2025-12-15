from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context
import os
import json
import zipfile
import io
import shutil
import secrets
import string
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from decimal import Decimal, InvalidOperation

app = Flask(__name__)

AUTOFILL_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'static', 'resources', 'mal.xlsx')


def _load_autofill_template_workbook():
    if not os.path.exists(AUTOFILL_TEMPLATE_PATH):
        raise FileNotFoundError(f"Fant ikke malfilen: {AUTOFILL_TEMPLATE_PATH}")
    return load_workbook(AUTOFILL_TEMPLATE_PATH, data_only=False)


def _get_basic_sheet(wb):
    if 'Basic' not in wb.sheetnames:
        raise KeyError("Fant ikke arket 'Basic' i malen.")
    return wb['Basic']


def _find_first_empty_row(ws, check_col_letter='A', start_row=2, max_scan_rows=1000):
    col = ws[check_col_letter]
    max_row = min(ws.max_row or 0, max_scan_rows)
    for r in range(start_row, max_row + 1):
        if ws[f'{check_col_letter}{r}'].value in (None, ''):
            return r
    return max_row + 1


def _fill_basic_row(ws, row_idx, firstname, hl_code, number, password):
    email = f"{number}@sikt.sykehuspartner.no"
    combined = f"{firstname}, {hl_code}".strip().strip(',')

    ws[f'A{row_idx}'] = firstname
    ws[f'B{row_idx}'] = firstname
    ws[f'C{row_idx}'] = hl_code
    ws[f'D{row_idx}'] = hl_code
    ws[f'E{row_idx}'] = email
    ws[f'H{row_idx}'] = combined
    ws[f'I{row_idx}'] = combined
    ws[f'K{row_idx}'] = password
    ws[f'M{row_idx}'] = 'Helselogistikk plastikknummer'
    ws[f'P{row_idx}'] = 'da'
    ws[f'Q{row_idx}'] = '(+2:0)Amsterdam, Berlin, Rome, Belgrade, Prague, Brussels, Sarajevo'
    ws[f'T{row_idx}'] = 'UNASSIGNED'
    ws[f'U{row_idx}'] = True


def normalize_imei(value):
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        s = str(value).strip()
        return s if len(s) == 15 and s.isdigit() else None

    if isinstance(value, float):
        if not value.is_integer():
            return None
        s = str(int(value)).strip()
        return s if len(s) == 15 and s.isdigit() else None

    s = str(value).strip()
    if s.endswith('.0'):
        s = s[:-2]

    if len(s) == 15 and s.isdigit():
        return s

    try:
        d = Decimal(s)
        if d != d.to_integral_value():
            return None
        s2 = str(d.to_integral_value())
        return s2 if len(s2) == 15 and s2.isdigit() else None
    except (InvalidOperation, ValueError):
        return None


def generate_secure_password(length=10):
    lowercase = string.ascii_lowercase
    uppercase = string.ascii_uppercase
    digits = string.digits

    password = [
        secrets.choice(lowercase),
        secrets.choice(uppercase),
        secrets.choice(digits)
    ]
    
    # Fill the rest with random characters from all sets
    all_chars = lowercase + uppercase + digits
    for _ in range(length - len(password)):
        password.append(secrets.choice(all_chars))
    
    # Shuffle to avoid predictable patterns
    secrets.SystemRandom().shuffle(password)
    return ''.join(password)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/generate_single', methods=['POST'])
def generate_single_file():
    if request.content_type and 'multipart/form-data' in request.content_type:
        code = request.form.get('code', '').strip().lower()
        imei_raw = request.form.get('imei', '')
        phone_raw = request.form.get('phone', '')
        firstname_raw = request.form.get('firstname', '')
        name_raw = request.form.get('name', '')
        lastname_raw = request.form.get('lastname', '')
    else:
        data = request.get_json() or {}
        code = str(data.get('code', '')).strip().lower()
        imei_raw = data.get('imei')
        phone_raw = data.get('phone')
        firstname_raw = data.get('firstname')
        name_raw = data.get('name')
        lastname_raw = data.get('lastname')

    if not code:
        return jsonify({'error': 'Code er påkrevd.'}), 400

    imei = normalize_imei(imei_raw)
    if not imei:
        return jsonify({'error': 'Ugyldig IMEI. IMEI må være nøyaktig 15 siffer.'}), 400

    phone = str(phone_raw).strip() if phone_raw is not None else ''
    if not phone:
        return jsonify({'error': 'Telefonnummer er påkrevd.'}), 400

    firstname_value = firstname_raw if firstname_raw is not None and str(firstname_raw).strip() else name_raw
    firstname = str(firstname_value).strip() if firstname_value is not None else ''
    if not firstname:
        return jsonify({'error': 'Fornavn er påkrevd.'}), 400

    lastname = str(lastname_raw).strip() if lastname_raw is not None else ''
    if not lastname:
        return jsonify({'error': 'Etternavn er påkrevd.'}), 400

    shutil.rmtree('avaya', ignore_errors=True)
    shutil.rmtree('ascom', ignore_errors=True)
    os.makedirs('avaya', exist_ok=True)
    os.makedirs('ascom', exist_ok=True)

    password = generate_secure_password(10)

    phn_content = f"SET SIPUSERNAME {phone}\nSET SIPUSERPASSWORD {password}\nGET /mdm/{code}/avaya/rw-sikt.txt"
    phn_filename = f"{imei}.phn"
    with open(f"avaya/{phn_filename}", 'w') as f:
        f.write(phn_content)

    json_content = {"voip_device_id": phone}
    json_filename = f"{imei}.json"
    with open(f"ascom/{json_filename}", 'w') as f:
        json.dump(json_content, f, indent=2)

    try:
        output_wb = _load_autofill_template_workbook()
        basic_ws = _get_basic_sheet(output_wb)
        start_row = 11
        _fill_basic_row(
            basic_ws,
            start_row,
            firstname=firstname,
            hl_code=lastname,
            number=phone,
            password=password
        )
    except Exception as e:
        return jsonify({'error': f'Kunne ikke bruke Excel-mal: {str(e)}'}), 400

    timestamp = datetime.now().strftime('%Y%m%d%H%M')
    temp_zip = f"temp_single_{secrets.token_hex(4)}.zip"
    download_zip_name = f"{code}_{phone}_{timestamp}.zip"
    output_xlsx_path = download_zip_name.replace('.zip', '.xlsx')
    output_wb.save(output_xlsx_path)

    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w') as zf:
        for root, dirs, files in os.walk('avaya'):
            for file in files:
                zf.write(os.path.join(root, file), os.path.join('avaya', file))
        for root, dirs, files in os.walk('ascom'):
            for file in files:
                zf.write(os.path.join(root, file), os.path.join('ascom', file))
        zf.write(output_xlsx_path, output_xlsx_path)

    memory_file.seek(0)
    with open(temp_zip, 'wb') as f:
        f.write(memory_file.getvalue())

    shutil.rmtree('avaya', ignore_errors=True)
    shutil.rmtree('ascom', ignore_errors=True)
    if os.path.exists(output_xlsx_path):
        os.remove(output_xlsx_path)

    return jsonify({
        'download_url': f"/download/{temp_zip}?download_name={download_zip_name}",
        'filename': download_zip_name
    })


@app.route('/generate', methods=['POST'])
def generate_files():
    # Handle both FormData and JSON
    if request.content_type and 'multipart/form-data' in request.content_type:
        code = request.form.get('code', '').strip().lower()
        try:
            start_num = int(request.form.get('start', 0))
            end_num = int(request.form.get('end', 0))
        except (ValueError, TypeError):
            return jsonify({'error': 'Ugyldig start- eller sluttnummer.'}), 400
        role_file = request.files.get('roleFile')
    else:
        data = request.get_json()
        code = data.get('code', '').strip().lower()
        try:
            start_num = int(data.get('start', 0))
            end_num = int(data.get('end', 0))
        except (ValueError, TypeError):
            return jsonify({'error': 'Ugyldig start- eller sluttnummer.'}), 400
        role_file = None
    
    # Validate input
    if role_file is None:
        if not code or start_num > end_num or not code.strip():
            return jsonify({'error': 'Vennligst fyll ut alle feltene korrekt. Kode er påkrevd.'}), 400
    
    # Read import xlsx if provided
    imeis = []
    role_names = []
    imported_codes = []
    if role_file:
        try:
            wb = load_workbook(role_file)
            ws = wb.active
            for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=3, values_only=True), start=1):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue

                imei_raw = row[0]
                name_raw = row[1] if len(row) > 1 else ''
                code_raw = row[2] if len(row) > 2 else ''

                imei = normalize_imei(imei_raw)
                if not imei:
                    return jsonify({'error': f'Ugyldig IMEI i kolonne A på rad {row_idx}. IMEI må være nøyaktig 15 siffer.'}), 400

                imported_code = str(code_raw).strip().lower() if code_raw is not None else ''
                if not imported_code:
                    return jsonify({'error': f'Mangler code i kolonne C på rad {row_idx}.'}), 400

                imeis.append(imei)
                role_names.append(str(name_raw).strip() if name_raw is not None else '')
                imported_codes.append(imported_code)
        except Exception as e:
            return jsonify({'error': f'Kunne ikke lese xlsx-fil: {str(e)}'}), 400
    
    # Clear previous files
    shutil.rmtree('avaya', ignore_errors=True)
    shutil.rmtree('ascom', ignore_errors=True)
    os.makedirs('avaya', exist_ok=True)
    os.makedirs('ascom', exist_ok=True)
    
    if role_file:
        total_files = len(imeis)
        temp_zip = f"temp_import_{secrets.token_hex(4)}.zip"
    else:
        total_files = end_num - start_num + 1
        temp_zip = f"temp_{code}_{start_num}_{end_num}.zip"

    timestamp = datetime.now().strftime('%Y%m%d%H%M')
    download_zip_name = f"{code if code else 'import'}_{start_num}_{end_num}_{timestamp}.zip"
    
    # Store data for output xlsx
    output_data = []
    
    def generate():
        # Generate files
        if role_file:
            for i, (imei, role_name, row_code) in enumerate(zip(imeis, role_names, imported_codes), 1):
                password = generate_secure_password(10)

                phone_number = start_num + (i - 1)

                output_data.append({
                    'role_name': role_name,
                    'hl_code': row_code,
                    'number': phone_number,
                    'password': password
                })

                phn_content = f"SET SIPUSERNAME {phone_number}\nSET SIPUSERPASSWORD {password}\nGET /mdm/{row_code}/avaya/rw-sikt.txt"
                phn_filename = f"{imei}.phn"
                with open(f"avaya/{phn_filename}", 'w') as f:
                    f.write(phn_content)

                json_content = {"voip_device_id": str(phone_number)}
                json_filename = f"{imei}.json"
                with open(f"ascom/{json_filename}", 'w') as f:
                    json.dump(json_content, f, indent=2)

                if i % 10 == 0 or i == total_files:
                    progress = int((i / total_files) * 100) if total_files else 100
                    yield f"data: {json.dumps({'progress': progress})}\n\n"
        else:
            for i, number in enumerate(range(start_num, end_num + 1), 1):
                # Generate cryptographically secure password
                password = generate_secure_password(10)
                
                # Get role name and HL code if available
                role_name = role_names[i-1] if i-1 < len(role_names) else ''
                hl_code = f'HL {code.upper()}'
                
                # Store data for output xlsx
                output_data.append({
                    'role_name': role_name,
                    'hl_code': hl_code,
                    'number': number,
                    'password': password
                })
                
                # Generate .phn file with the secure password
                phn_content = f"SET SIPUSERNAME {number}\nSET SIPUSERPASSWORD {password}\nGET /mdm/{code}/avaya/rw-sikt.txt"
                phn_filename = f"{code}{number}.phn"
                with open(f"avaya/{phn_filename}", 'w') as f:
                    f.write(phn_content)
                
                # Generate .json file
                full_value = f"{number}"
                json_content = {"voip_device_id": full_value}
                json_filename = f"{full_value}.json"
                with open(f"ascom/{json_filename}", 'w') as f:
                    json.dump(json_content, f, indent=2)
                
                # Calculate and send progress (only send every 10 files or at completion to reduce overhead)
                if i % 10 == 0 or i == total_files:
                    progress = int((i / total_files) * 100)
                    yield f"data: {json.dumps({'progress': progress})}\n\n"
        
        # Create output xlsx from template and fill only Basic sheet
        try:
            output_wb = _load_autofill_template_workbook()
            basic_ws = _get_basic_sheet(output_wb)
            start_row = 11

            for offset, data_row in enumerate(output_data):
                _fill_basic_row(
                    basic_ws,
                    start_row + offset,
                    firstname=data_row.get('role_name', ''),
                    hl_code=data_row.get('hl_code', ''),
                    number=data_row.get('number', ''),
                    password=data_row.get('password', '')
                )
        except Exception as e:
            yield f"data: {json.dumps({'error': f'Kunne ikke bruke Excel-mal: {str(e)}'})}\n\n"
            return
        
        # Save output xlsx
        output_xlsx_path = download_zip_name.replace('.zip', '.xlsx')
        output_wb.save(output_xlsx_path)
        
        # Create zip file
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w') as zf:
            # Add avaya files
            for root, dirs, files in os.walk('avaya'):
                for file in files:
                    zf.write(os.path.join(root, file), 
                            os.path.join('avaya', file))
            
            # Add ascom files
            for root, dirs, files in os.walk('ascom'):
                for file in files:
                    zf.write(os.path.join(root, file), 
                            os.path.join('ascom', file))
            
            # Add output xlsx to root of zip
            zf.write(output_xlsx_path, output_xlsx_path)
        
        memory_file.seek(0)
        
        # Save zip to a temporary file
        with open(temp_zip, 'wb') as f:
            f.write(memory_file.getvalue())
        
        # Clean up
        shutil.rmtree('avaya', ignore_errors=True)
        shutil.rmtree('ascom', ignore_errors=True)
        if os.path.exists(output_xlsx_path):
            os.remove(output_xlsx_path)
        
        # Send completion event
        if role_file:
            yield f"data: {json.dumps({'complete': True, 'download_url': f'/download/{temp_zip}?download_name={download_zip_name}', 'filename': download_zip_name})}\n\n"
        else:
            yield f"data: {json.dumps({'complete': True, 'download_url': f'/download/{temp_zip}?download_name={download_zip_name}', 'filename': download_zip_name})}\n\n"
    
    return Response(stream_with_context(generate()), mimetype='text/event-stream')

@app.route('/download/<filename>')
def download_file(filename):
    # Security check: only allow temp_ prefixed files
    if not filename.startswith('temp_') or '..' in filename:
        return jsonify({'error': 'Invalid filename'}), 400

    download_name = request.args.get('download_name')
    if download_name:
        download_name = download_name.strip()
        if not download_name.lower().endswith('.zip'):
            download_name += '.zip'
        if (
            '..' in download_name
            or '/' in download_name
            or '\\' in download_name
            or not re.fullmatch(r'[A-Za-z0-9._-]+', download_name)
        ):
            return jsonify({'error': 'Invalid download name'}), 400
    else:
        download_name = filename.replace('temp_', '')
    
    response = send_file(
        filename,
        as_attachment=True,
        download_name=download_name
    )
    
    # Clean up temp file after sending
    @response.call_on_close
    def cleanup():
        try:
            if os.path.exists(filename):
                os.remove(filename)
        except Exception:
            pass
    
    return response

if __name__ == '__main__':
    app.run(debug=True)